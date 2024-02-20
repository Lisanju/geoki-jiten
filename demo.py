import openpyxl

# Path do workbook okinawano
wbOK = openpyxl.load_workbook("data/ok_okinawago_data.xlsx")
# Worksheet okinawano ativo
wsOK = wbOK.active

# Path do workbook português brasileiro
wbPB = openpyxl.load_workbook("data/ptbr_okinawago_data.xlsx")
#Worksheet português brasileiro ativo
wsPB = wbPB.active

# Função para pesquisar palavra no worksheet, retornando sua célula e informações lexicais
def acharPalavra(string):
    for i in range(1,wsOK.max_row+1):
        for j in range(1,2):
            if string == wsOK.cell(i,j).value:
                celula = wsOK.cell(i,j)
                palavra = wsPB.cell(i,j).value # FIX: Arrumar palavra
                entonacao = wsOK.cell(i,j+1).value
                classe = wsOK.cell(i,j+2).value
                glosa = wsPB.cell(i,j+2).value # FIX: Arrumar glosa
                print(f"Palavra encontrada: {wsOK.cell(i,j).value}\nEntonação: {entonacao}\nClasse: {classe}")
                return celula, palavra, entonacao, classe, glosa
                
acharPalavra("?aci")

