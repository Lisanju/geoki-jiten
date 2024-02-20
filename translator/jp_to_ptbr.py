import openpyxl
import translators as ts
import time

# Path do workbook português brasileiro (ainda não traduzido do JP para PT-BR)
wb = openpyxl.load_workbook("../data/ptbr_okinawago_data.xlsx")

# Worksheet português brasileiro ativo
ws = wb.active

# Função com loop de tradução que itera sobre cada célula da planilha
# Caso a célula esteja vazia ou não tenha uma string como valor, não há tradução
# Há um intervalo de 1 segundo para cada tradução
def tradutor():
  for i in range(1, ws.max_row + 1):
      for j in range(1, ws.max_column + 1):
          celulaValor = ws.cell(i, j).value
          if celulaValor and isinstance(celulaValor, str):
              palavraTraduzida = ts.translate_text(celulaValor, "alibaba")
              ws.cell(row = i, column = j, value = palavraTraduzida)
              time.sleep(1)

tradutor()

# Salvar o workbook traduzido
wb.save("ptbr_okinawago_data.xlsx")
