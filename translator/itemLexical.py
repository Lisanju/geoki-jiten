import openpyxl
import translators as ts
import time

wb = openpyxl.load_workbook("data/jp_okinawago_data.xlsx")
ws = wb.active

linhasProcessadas = 0

for i in range(1, ws.max_column+1):
    for j in range(1, ws.max_column + 1):
        celulaValor = ws.cell(i, j).value
        if celulaValor and isinstance(celulaValor, str):
            try:
                palavraTraduzida = ts.translate_text(celulaValor, "baidu","jp","pt")
                if palavraTraduzida is not None:
                    ws.cell(row=i, column=j, value=palavraTraduzida)
                    print(f"{i}/{ws.max_row}")
                else:
                    print(f"A tradução para a célula ({i}, {j}) retornou None.")
            except Exception as e:
                print(f"Ocorreu um erro ao traduzir a célula ({i}, {j}): {e}")
            time.sleep(1)
    linhasProcessadas = linhasProcessadas + 1
    if linhasProcessadas % 100 == 0 or i == ws.max_row:
        wb.save("data/ptbr_okinawago_data.xlsx")
        print("Arquivo salvo.")

wb.save("ptbr_okinawago_data.xlsx")