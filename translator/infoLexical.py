import openpyxl
import re

# Path do workbook japonês
wb = openpyxl.load_workbook("../data/jp_okinawago_data.xlsx")
# Worksheet japonês ativo
ws = wb.active

# Criar uma lista para cada coluna
palavra = [cell.value for cell in ws["A"]]
entonacao = [cell.value for cell in ws["B"]]
classe = [cell.value for cell in ws["C"]]

# Traduzir as classes do japonês para o português brasileiro
def traduzirClasse(classeJapones):
    substituicoes = {
        "名": "Nome ",
        "感": "Interjeição ",
        "副": "Advérbio ",
        "接続": "Afixo ",
        "接頭": "Prefixo ",
        "助": "Partícula ",
        "接尾": "Sufixo ",
        "連体": "Adjunto adnominal ",
        "他": "Verbo transitivo ",
        "助詞": "Partícula gramatical ",
        "動": "Verbo ",
        "助詞的に用いられる": "Usado como partícula ",
        "句": "Expressão",
        "各": "Onomatopeia",
        "形": "Adjetivo ",
        "自": "Verbo intransitivo ",
        "連詞": "Conjunção ",
        "不規則": "Irregular ",
        "文": "Formulaico ",
        "一部": "Grupo de ",
        "否定": "Negação de ",
        "のみがある": "",
        "活用は": "Função igual a ",
        "詞的に用いられる": ""
    }
    return substituicoes.get(classeJapones, classeJapones)

classeTraduzida = [mapear_classe(classeJapones) for classeJapones in classe]

# Criar uma lista de listas com os dados traduzidos, mantendo entonação e palavra sem modificação
dadosTraduzidos = list(zip(palavra, entonacao, classe_traduzida))

# Exportar para um novo arquivo Excel com os dados traduzidos
wbTraduzido = openpyxl.Workbook()
wsTraduzido = wb_traduzido.active

# Escrever os dados traduzidos no arquivo Excel
for row_idx, row_data in enumerate(dadosTraduzidos, start=1):
    for col_idx, cell_value in enumerate(row_data, start=1):
        wsTraduzido.cell(row=row_idx, column=col_idx, value=cell_value)

# Salvar o arquivo Excel traduzido
wbTraduzido.save("../data/ok_okinawago_data.xlsx")
